"""Sanity-check the exported Excel."""
import sys
from openpyxl import load_workbook

path = sys.argv[1]
wb = load_workbook(path)
print(f"Sheets: {wb.sheetnames}\n")

for name in wb.sheetnames:
    ws = wb[name]
    print(f"=== {name} ({ws.max_row} rows × {ws.max_column} cols) ===")
    # Print first 4 rows
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 4: break
        cells = [str(c)[:40] if c is not None else '' for c in row[:8]]
        print(f"  Row {i+1}: {cells}")
    print()
